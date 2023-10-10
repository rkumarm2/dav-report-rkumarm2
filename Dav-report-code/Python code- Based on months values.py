#!/usr/bin/env python
# coding: utf-8

# In[1]:


#pip install tk
#pip install xlsxwriter
#pip install xlrd
#pip install csv
#pip install pandas
#pip install csv
#pip install openpyxl


# In[]:


import pandas as pd
import tk
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill

from tkinter import ttk
from tkinter import *
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfiles

import csv
from operator import itemgetter
import os
import glob
import time
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font
import xlrd as xl


# In[6]:


root=tk.Tk()

canvas = tk.Canvas(root,width=350,height = 450)
canvas.grid(columnspan=3,rowspan=12)
logo = Image.open('logo.png')
logo = ImageTk.PhotoImage(logo)
logo_label=tk.Label(image=logo)
logo_label.image =logo
logo_label.grid(column=1,row=0)

instructions = tk.Label(root, text ="Enter today's date in yyyy/mm/dd format :",font="Raleway")
instructions.grid(columnspan=3,column=0,row=1)

datevalue = StringVar()
dateentry = Entry(root, textvariable=datevalue)
dateentry.grid(columnspan=3,column=0,row=2)

#
instructions = tk.Label(root, text ="  ",font="Raleway")
instructions.grid(columnspan=3,column=0,row=3)
instructions = tk.Label(root, text ="Enter difference in months : E.g. Aug - May = 8-5=3",font="Raleway")
instructions.grid(columnspan=3,column=0,row=4)

the_delta = IntVar()
deltaentry = Entry(root, textvariable=the_delta)
deltaentry.grid(columnspan=3,column=0,row=5)
#

instructions = tk.Label(root, text =" ",font="Raleway")
instructions.grid(columnspan=3,column=0,row=6)

instructions = tk.Label(root, text ="Select the files",font="Raleway")
instructions.grid(columnspan=3,column=0,row=7)

#browse button

browse_text = tk.StringVar()
browse_btn = tk.Button(root, textvariable=browse_text, command=lambda:open_file(),
                       font ="Raleway", bg="green", fg="white",height=2,width=12)
browse_text.set("Browse")
browse_btn.grid(columnspan=3,column=1,row=8)

instructions2 = tk.Label(root, text ="  ",font="Raleway")
instructions2.grid(columnspan=3,column=0,row=9)

my_progress = ttk.Progressbar(root, orient = HORIZONTAL, length = 400, mode='determinate')
my_progress.grid(columnspan=3,column=0,row=10)

instructions3 = tk.Label(root, text ="  ",font="Raleway")
instructions3.grid(columnspan=3,column=0,row=11)



# In[ ]:


Month_list=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
value_list=["01","02","03","04","05","06","07","08","09","10","11","12"]

def open_file():
    given_delta = the_delta.get()
    dx = datevalue.get()
    d2 = datetime.strptime(dx, "%Y/%m/%d")
    curyear , curmon , curdate = dx.split("/")
    curmon , curyear = int(curmon) , int(curyear)
    browse_text.set("loading...")
    
    file = askopenfiles(parent=root, mode='rb', title="`Choose a file")
    #list for all the devices
    csvfile_list=[]
    #list for view managed device list 
    csvfile_vmd_list=[]
    #list for detailed device details list
    csvfile_ddd_list=[]
        
    for i in file:
        csvfile_list.append(i.name)
        
    for i in csvfile_list:
        t= i.split("/")
        if(t[-1][5]=="V"):
            csvfile_vmd_list.append(i)
        else:
            csvfile_ddd_list.append(i)
    csvfile_vmd_list.sort()
    csvfile_ddd_list.sort()
    date = csvfile_list[0].split("_")[-1].split(".")[0]
 
    xlsxfile_list=[]
    testcsvfile_list=[]
    file_number=0
    status_coloumn_list=[]
    ip_address_list=[]
    
#working on detailed devices list    
    for i in csvfile_ddd_list:
        f = open(i)
        data = list(csv.reader(f))
        f.close()
        lst = data
        lst = sorted(data, key = itemgetter(2))                              
        header = ["    Ip Address  ", "   PRIMARYDEVICENAME   ","   STATUS"   , "   SSHV2_STATUS   ","   TELNET_STATUS   ", "   SNMPV2C_STATUS   "]
        newdetailedfile = "test" + i
        newdetailedfile = newdetailedfile.split("/")
        newdetailedfile = newdetailedfile[-1]
        newdetailedfile = "test" + newdetailedfile

        my_progress['value']+=9
        root.update_idletasks()

        status_coloumn=[]
        ip_address=[]

        for i in range(0,len(lst)-1):
            gtime=lst[i][17]
            if gtime=="" or gtime==" " or gtime==None:
                gtime_avl=0
            else:
                gtime_avl=1
                gtime=gtime.split(" ")
            if(gtime_avl == 1):
                gd , gm , gy = gtime[2] , gtime[1], int(gtime[-1])
                gm=int(value_list[Month_list.index(gm)])
                if curyear>gy : themon = curmon + 12
                else: themon = curmon
                delta = themon - gm

            ip_address.append(lst[i][2])
            statval=lst[i][15]
                
            if(lst[i][24] != "Successful"):
                statval = "SNMP Failed"
      
            if(lst[i][24] == "Successful" and lst[i][15]!="Successful" and lst[i][18]== "Successful"):
                statval = "Telnet Successful"
      
            if statval == "" :
                statval = "Connection Failed"

            if ((gtime_avl and (  delta >= (given_delta +1) ) ) or gtime_avl!=1):
                statval = "Connection Failed"            
                
            status_coloumn.append(statval)
            
        status_coloumn_list.append(status_coloumn)
        ip_address_list.append(ip_address)
        f.close()
      
#working on view managed devices list
    
    for i in csvfile_vmd_list:
        my_progress['value']+=9
        root.update_idletasks()
        
        f = open(i)
        data = list(csv.reader(f))
        f.close()
        lst = data

        lst.sort()
        header = ["    Ip Address    ", "    Display Name    ","    STATUS    ", "    Device Family    ","    Product Model    ", "    Os Name    ","    Os Version    "]
        newfile = "test" + i
        newfile = newfile.split("/")
        newfile = newfile[-1]
        newfile = "test" + newfile
        
        f = open(newfile ,"w+",newline="")
       
        obj = csv.writer(f)
        signal =1
        ip_error_name=["Note: IP address is not matching    "]
        count_error_name=["Note: Device count is not matching    "]
        
        if(len(lst)!=len(status_coloumn_list[file_number])+1):
            signal=0
            obj.writerow([""])
            obj.writerow(["Error"])
            obj.writerow(count_error_name)
            
        else:
            obj.writerow(header)

        if(signal!=0):
            for i in range(0,len(lst)-1):
                #condition for checking ip adrress pf vmd file and ddd file
                if(lst[i][0]!=ip_address_list[file_number][i] ):
                    f.truncate(0)
                    obj.writerow(["Error"])
                    obj.writerow(["Error"])
                    
                    ip_error="Note: Ip address - " + lst[i][0] + "(VMD) and " +ip_address_list[file_number][i] + "(DDD) is not matching."
                    obj.writerow([ip_error])
                    #print(ip_error)
                    break
                else:
                    obj.writerow([lst[i][0], lst[i][2],status_coloumn_list[file_number][i], lst[i][4], lst[i][5], lst[i][8], lst[i][9]])
                    
                
                #condion end
        file_number+=1    
        f.close()
        
        read_file = pd.read_csv (newfile)
        testcsvfile_list.append(newfile)
        newfile= newfile.split(".")[0]
        newxlsxfile = newfile + ".xlsx"
        
        xlsxfile_list.append(newxlsxfile)
        read_file.to_excel (newxlsxfile, index = None, header=True)

        
    ws_name=[]
    for i in xlsxfile_list:
        ws_name.append(i[4:8])

    final_file_name = "Charter_Spectrum_Enterprise_DAV_Report_" + date + ".xlsx"
    #print(final_file_name)
    writer = pd.ExcelWriter(final_file_name)
    #print(xlsxfile_list)

    for excel_file in xlsxfile_list:
        my_progress['value']+=9
        root.update_idletasks()

        sheet=excel_file
        sheet = sheet[4:8]
        df1=pd.read_excel(excel_file)
        df1.fillna(value="", inplace=True)
        df1.to_excel(writer ,sheet_name= sheet, index=False)

        #for adjusting coloum width 
        for column in df1:
            column_width = max(df1[column].astype(str).map(len).max(), len(column))
            col_idx = df1.columns.get_loc(column)
            writer.sheets[sheet].set_column(col_idx, col_idx, column_width)

        #for making filter in column 
        (max_row, max_col) = df1.shape            
        writer.sheets[sheet].autofilter(0, 0, max_row, max_col - 1)
        writer.sheets[sheet].filter_column(0, 'Region == East')

        #for coloring
        def color_negative_red(val):
            """
            Takes a scalar and returns a string with
            the css property `'color: red'` for negative
            strings, black otherwise.
            """
            color = 'red' 
            return 'color: %s' % color
    
    writer.save()

    browse_text.set("Browse")
    #instructions
        
    instructions = tk.Label(root, text ="File is downloaded in current directory",font="Raleway")

    

    #for removing extra file
    for i in xlsxfile_list:
        os.remove(i)
    for i in testcsvfile_list:
        os.remove(i)

    wb =openpyxl.load_workbook(final_file_name) 
    blue_color='8ea9db'
    red_color='FF7F7F'

    for i in ws_name:
        my_progress['value']+=6.5
        root.update_idletasks()
        #progressbarclose
        ws=wb[i]
        #print(i)
        value_of_sheet = ws['A2'].value
        #print(value_of_sheet)
        if(value_of_sheet!="Error"):
            fill_pattern = PatternFill(patternType='solid',fgColor=blue_color)
            c=['A1','B1','C1','D1','E1','F1','G1']
            for y in c:
                ws[y].fill=fill_pattern
        else:
            ws['A1']="Not able to process the file.    "
            fill_pattern = PatternFill(patternType='solid',fgColor=red_color)   
            ws['A2'].fill=fill_pattern
              
    wb.save(final_file_name)
    wb = load_workbook(final_file_name)
    ws1 = wb['SEBH']
    ws2 = wb['SECH']
    ws3 = wb['SETW']
    
    for ws in [ws1 , ws2 , ws3]:
        rc = ws.max_row
        cc = ws.max_column  
        #print(ws , rc , cc)
        
        for i in range(1, rc+1 ):
            cell="C"+str(i)
            if ws[cell].value == "Connection Failed":
                ws[cell].font = Font(b=True)
    #print("Bold Conversion Completed")
    wb.save(final_file_name)
    instructions = tk.Label(root, text ="File is downloaded in current directory",font="Raleway")
    instructions.grid(columnspan=3,column=0,row=9)
    
root.mainloop()